# Lollms function call definition file
from functools import partial
from lollms.utilities import PackageManager
from ascii_colors import trace_exception

# Install openpyxl if not already installed
if not PackageManager.check_package_installed("openpyxl"):
    PackageManager.install_package("openpyxl")

import openpyxl

# Function to add a training session
def add_training_session(filepath: str, week: int, volume: int, intensity: int, peaking: int) -> str:
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        new_row = [week, volume, intensity, peaking]
        ws.append(new_row)
        wb.save(filepath)
        return f"Training session for week {week} added successfully!"
    except Exception as e:
        return trace_exception(e)

# Metadata function for add_training_session
def add_training_session_function(filepath):
    return {
        "function_name": "add_training_session",
        "function": partial(add_training_session, filepath=filepath),
        "function_description": "Adds a training session to the Excel file.",
        "function_parameters": [
            {"name": "week", "type": "int"},
            {"name": "volume", "type": "int"},
            {"name": "intensity", "type": "int"},
            {"name": "peaking", "type": "int"}
        ]
    }

# Function to update a training session
def update_training_session(filepath: str, week: int, volume: int, intensity: int, peaking: int) -> str:
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
            if row[0].value == week:
                row[1].value = volume
                row[2].value = intensity
                row[3].value = peaking
                wb.save(filepath)
                return f"Training session for week {week} updated successfully!"
        return f"Week {week} not found in the training schedule."
    except Exception as e:
        return trace_exception(e)

# Metadata function for update_training_session
def update_training_session_function():
    return {
        "function_name": "update_training_session",
        "function": partial(update_training_session, filepath=filepath)
        "function_description": "Updates a training session in the Excel file.",
        "function_parameters": [
            {"name": "week", "type": "int"},
            {"name": "volume", "type": "int"},
            {"name": "intensity", "type": "int"},
            {"name": "peaking", "type": "int"}
        ]
    }

# Function to delete a training session
def delete_training_session(filepath: str, week: int) -> str:
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
            if row[0].value == week:
                ws.delete_rows(row[0].row, 1)
                wb.save(filepath)
                return f"Training session for week {week} deleted successfully!"
        return f"Week {week} not found in the training schedule."
    except Exception as e:
        return trace_exception(e)

# Metadata function for delete_training_session
def delete_training_session_function():
    return {
        "function_name": "delete_training_session",
        "function": partial(delete_training_session, filepath=filepath)
        "function_description": "Deletes a training session from the Excel file.",
        "function_parameters": [
            {"name": "week", "type": "int"}
        ]
    }
